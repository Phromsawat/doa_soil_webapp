# -*- coding: utf-8 -*-
"""Import the DOA fixed fertilizer-plan tables (E:\\DOA_Soil\\คำแนะนำ) into SQL.

Each source sheet is a *pivoted* lookup table:
  row with label 'อินทรียวัตถุ'  -> OM range per data column
  row with label 'ฟอสฟอรัส'      -> P range  per data column
  row with label 'โพแทสเซียม'     -> K range  per data column
  following rows -> (stage, grade) x amount per soil-combo column
Output: supabase/migrations/008_import_fertilizer_plan.sql
"""
import openpyxl, os, re, sys
sys.stdout.reconfigure(encoding="utf-8")

FOLDER = r"E:\DOA_Soil\คำแนะนำ"
OUT = r"D:\doa_test_kit\supabase\migrations\008_import_fertilizer_plan.sql"

# (file, sheet, crop_name in our DB, use_type)
MAP = [
    ("1. อ้อย-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "อ้อยปลูก-แม่ปุ๋ย",        "อ้อยปลูก",          "straight"),
    ("1. อ้อย-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "อ้อยปลูก-ปุ๋ยเชิงประกอบ",  "อ้อยปลูก",          "compound"),
    ("1. อ้อย-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "อ้อยตอ-แม่ปุ๋ย",          "อ้อยตอ",            "straight"),
    ("1. อ้อย-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "อ้อยตอ-ปุ๋ยเชิงประกอบ",    "อ้อยตอ",            "compound"),
    ("2. มันสำปะหลัง-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "มันสำปะหลัง-แม่ปุ๋ย",       "มันสำปะหลัง",       "straight"),
    ("2. มันสำปะหลัง-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "มันสำปะหลัง-ปุ๋ยเชิงประกอบ", "มันสำปะหลัง",       "compound"),
    ("3. ข้าวโพด-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ข้าวโพดเลี้ยงสัตว์-แม่ปุ๋ย",  "ข้าวโพดเลี้ยงสัตว์", "straight"),
    ("3. ข้าวโพด-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ข้าวโพดเลี้ยงสัตว์-ปุ๋ยเชิงประก", "ข้าวโพดเลี้ยงสัตว์", "compound"),
    ("3. ข้าวโพด-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ข้าวโพดฝักสด-แม่ปุ๋ย",       "ข้าวโพดฝักสด",      "straight"),
    ("3. ข้าวโพด-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ข้าวโพดฝักสด-ปุ๋ยเชิงประกอบ", "ข้าวโพดฝักสด",      "compound"),
    ("4. ถั่ว-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ถั่ว-แม่ปุ๋ย",             "ถั่ว",              "straight"),
    ("4. ถั่ว-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", "ถั่ว-ปุ๋ยเชิงประกอบ",       "ถั่ว",              "compound"),
]
# ไม้ผล (file 5): แต่ละ sheet = 1 พืช, มีเฉพาะแม่ปุ๋ย
for fruit in ["ทุเรียน","มังคุด","เงาะ","มะม่วง","ลำไย","ลิ้นจี่","ส้ม","มะพร้าว","สับปะรด"]:
    MAP.append(("5. ไม้ผล-คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน-ใหม่.xlsx", fruit, fruit, "straight"))


def norm(v):
    return "" if v is None else str(v).strip()

def clean_stage(s):
    s = re.sub(r"\s+", " ", norm(s))
    s = re.sub(r"\s*\((?:กรัม/ต้น|กก\./ไร่)\)", "", s)  # ตัดหน่วยออกจากชื่อระยะ
    return s.strip()

def detect_unit(label):
    if "กรัม/ต้น" in label: return "กรัม/ต้น"
    if "กก./ไร่" in label or "กก./ไร" in label: return "กก./ไร่"
    return None

def parse_range(s):
    s = norm(s).replace(" ", "")
    if not s or s == "-":
        return (None, None)
    m = re.match(r"^<([\d.]+)$", s)
    if m: return (None, float(m.group(1)))
    m = re.match(r"^>([\d.]+)$", s)
    if m: return (float(m.group(1)), None)
    m = re.match(r"^([\d.]+)-([\d.]+)$", s)
    if m: return (float(m.group(1)), float(m.group(2)))
    raise ValueError(f"unparseable range: {s!r}")


def parse_sheet(path, sheet, crop, use_type):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    om_r = p_r = k_r = None
    for i, r in enumerate(rows[:8]):
        a = norm(r[0])
        if a.startswith("อินทรียวัตถุ"): om_r = i
        elif a.startswith("ฟอสฟอรัส"): p_r = i
        elif a.startswith("โพแทสเซียม"): k_r = i
    if None in (om_r, p_r, k_r):
        wb.close(); return None  # ไม่ใช่ตาราง lookup (เป็นข้อความ/หมายเหตุอย่างเดียว)
    ncol = ws.max_column
    combos = []
    for c in range(2, ncol):
        def cell(rr): return rows[rr][c] if c < len(rows[rr]) else None
        om, p, k = parse_range(cell(om_r)), parse_range(cell(p_r)), parse_range(cell(k_r))
        combos.append(None if (om == (None,None) and p == (None,None) and k == (None,None)) else (om, p, k))

    out, cur_stage, order, unit = [], None, 0, None
    for i in range(k_r + 1, len(rows)):
        r = rows[i]
        label = norm(r[0]); grade = norm(r[1]) if len(r) > 1 else ""
        if label.startswith("หมายเหตุ") or label.startswith("ที่มา"):
            break
        if label:
            u = detect_unit(label)
            if u: unit = u
            cur_stage = clean_stage(label); order += 1
        if not grade:
            continue
        for ci, combo in enumerate(combos):
            if combo is None: continue
            val = r[ci + 2] if (ci + 2) < len(r) else None
            if val is None or norm(val) == "": continue
            if not isinstance(val, (int, float)):  # skip stray text
                continue
            om, p, k = combo
            out.append((crop, use_type, om[0], om[1], p[0], p[1], k[0], k[1],
                        cur_stage, order, grade, float(val), unit))
    wb.close()
    if unit is None:  # fallback by crop type
        unit = "กก./ไร่"
        out = [t[:-1] + (unit,) for t in out]
    return out


def sql_num(x):
    return "NULL" if x is None else (str(int(x)) if float(x).is_integer() else str(x))

def sql_str(s):
    return "'" + norm(s).replace("'", "''") + "'"


def main():
    all_rows, summary, skipped = [], [], []
    for fname, sheet, crop, use_type in MAP:
        path = os.path.join(FOLDER, fname)
        rows = parse_sheet(path, sheet, crop, use_type)
        if rows is None:
            skipped.append((crop, use_type, sheet))
            continue
        all_rows.extend(rows)
        combos = len(set((r[2],r[3],r[4],r[5],r[6],r[7]) for r in rows))
        stages = len(set(r[8] for r in rows))
        unit = rows[0][12] if rows else "?"
        summary.append((crop, use_type, len(rows), combos, stages, unit))

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("-- 008_import_fertilizer_plan.sql — auto-generated by scripts/import_fertilizer_plan.py\n")
        f.write("-- ตารางแผนปุ๋ยตายตัวของกรมฯ (คำแนะนำการใช้ปุ๋ยตามค่าวิเคราะห์ดิน 2564) 5 ไฟล์\n")
        f.write(f"-- {len(all_rows)} rows across {len(MAP)} sheets\n\n")
        f.write("DELETE FROM public.crop_fertilizer_plan;\n\n")
        f.write("INSERT INTO public.crop_fertilizer_plan\n")
        f.write("  (crop_id, use_type, om_min, om_max, p_min, p_max, k_min, k_max, stage, stage_order, grade, amount, unit)\n")
        f.write("SELECT c.id, v.use_type, v.om_min, v.om_max, v.p_min, v.p_max, v.k_min, v.k_max, v.stage, v.stage_order, v.grade, v.amount, v.unit\n")
        f.write("FROM (VALUES\n")
        lines = []
        for idx, r in enumerate(all_rows):
            crop, use_type, om0, om1, p0, p1, k0, k1, stage, order, grade, amount, unit = r
            cast = "::numeric" if idx == 0 else ""
            vals = (f"({sql_str(crop)}, {sql_str(use_type)}, "
                    f"{sql_num(om0)}{cast}, {sql_num(om1)}{cast}, {sql_num(p0)}{cast}, {sql_num(p1)}{cast}, "
                    f"{sql_num(k0)}{cast}, {sql_num(k1)}{cast}, {sql_str(stage)}, {order}, {sql_str(grade)}, "
                    f"{sql_num(amount)}{cast}, {sql_str(unit)})")
            lines.append(vals)
        f.write(",\n".join(lines))
        f.write("\n) AS v(crop_name, use_type, om_min, om_max, p_min, p_max, k_min, k_max, stage, stage_order, grade, amount, unit)\n")
        f.write("JOIN public.crops c ON c.name = v.crop_name;\n")

    print(f"WROTE {OUT}")
    print(f"TOTAL rows: {len(all_rows)}\n")
    print(f"{'crop':<22}{'use_type':<10}{'rows':>6}{'combos':>8}{'stages':>8}  unit")
    for crop, ut, n, combos, stages, unit in summary:
        print(f"{crop:<22}{ut:<10}{n:>6}{combos:>8}{stages:>8}  {unit}")
    if skipped:
        print("\nSKIPPED (not a lookup table — text/note only):")
        for crop, ut, sheet in skipped:
            print(f"  {crop} [{ut}]  <- {sheet}")


if __name__ == "__main__":
    main()
