"""
สคริปต์สกัดตาราง KM ไม้ผล 100% จาก Excel ของกรมพัฒนาที่ดิน
ไปเป็น JSON ที่ src/lib/fruit_fertilizer.json

Usage:
    py scripts/extract_fruit_fertilizer.py

ต้องมี:
    pip install openpyxl
"""
import openpyxl, json, sys, io, pathlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "docs" / "references" / "100__ตารางคำนวณปุ๋ยสำหรับ KM ไม้ผล_04.08.2569.xlsx"
OUT_JSON = ROOT / "src" / "lib" / "fruit_fertilizer.json"

CROP_KEY = {
    "ทุเรียน100%": "durian",
    "มังคุด100%": "mangosteen",
    "เงาะ100%": "rambutan",
    "มะม่วง100%": "mango",
    "ลำไย100%": "longan",
    "ลิ้นจี่100%": "lychee",
    "ส้ม100%": "orange",
    "มะพร้าว100%": "coconut",
    "สับปะรด100%": "pineapple",
}

# เกณฑ์ Excel: OM (%): <2, 2-3, >3 | P (mg/kg): <15, 15-45, >45 | K: <50, 50-100, >100
LEVELS = ["low", "med", "high"]

# 4 ระยะ × 3 สูตรปุ๋ย (46-0-0 = urea, 18-46-0 = dap, 0-0-60 = kcl)
#
# ⚠️ คอลัมน์ของ "ระยะ" เริ่มที่ 26 ไม่ใช่ 23 — ยืนยันจากหัวตารางแถว 11 ของ Excel:
#     C23 = "**ปรับเป็นเลขจำนวนเต็มหลักร้อย"  ← คอลัมน์ผลรวมทั้งปี (ไม่ใช่ระยะ)
#     C26 = ระยะบำรุงต้น | C29 = ระยะสร้างตาดอก | C32 = ระยะบำรุงผล | C35 = ระยะปรับปรุงคุณภาพ
#   เดิมเริ่มที่ 23 ทำให้อ่าน "ผลรวม" มาเป็นระยะแรก แล้วทุกระยะเลื่อนไป 1 ระยะ
#   และระยะปรับปรุงคุณภาพหายไป -> ตัวเลขไม่ตรงกับเว็บกรมวิชาการเกษตร
STAGE_COL_START = 26
STAGES = ["nurture", "bud", "fruit", "quality"]

def main():
    wb = openpyxl.load_workbook(str(SRC_XLSX), data_only=True)
    data = {}
    for sheet_name, crop_key in CROP_KEY.items():
        ws = wb[sheet_name]
        crop_data = {}
        idx = 0
        for om in LEVELS:
            for p in LEVELS:
                for k in LEVELS:
                    row = 13 + idx  # rows 13-39 = 27 combos
                    key = f"{om}_{p}_{k}"
                    stage_vals = {}
                    for si, stage in enumerate(STAGES):
                        c_start = STAGE_COL_START + si * 3
                        stage_vals[stage] = {
                            "urea": int(ws.cell(row, c_start).value or 0),
                            "dap":  int(ws.cell(row, c_start + 1).value or 0),
                            "kcl":  int(ws.cell(row, c_start + 2).value or 0),
                        }
                    crop_data[key] = stage_vals
                    idx += 1
        data[crop_key] = crop_data

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_JSON}")
    print(f"Crops: {list(data.keys())}")

if __name__ == "__main__":
    main()
